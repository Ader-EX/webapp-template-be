from typing import Optional
from fastapi import APIRouter, Depends, HTTPException
from sqlalchemy.orm import Session,joinedload
from schemas import PaginatedResponseSchemas
from models.ConsManager import ConsultingManager
from models.ProjExperience import ProjectExperience
from schemas.ProjManagerSchema import ProjectExperienceCreate, ProjectExperienceResponse, ProjectExperienceUpdate
from database import get_db
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO
from datetime import datetime


router = APIRouter()



@router.post("/", response_model=ProjectExperienceResponse)
def create_experience(
    data: ProjectExperienceCreate,
    db: Session = Depends(get_db), 
):
    new_project = ProjectExperience(
        **data.dict(),
    )

    db.add(new_project)
    db.commit()
    db.refresh(new_project)
    return new_project


@router.get("/", response_model=PaginatedResponseSchemas.PaginatedResponse[ProjectExperienceResponse])
def get_experiences(
    skip: int = 0,
    limit: int = 10,
    search: Optional[str] = None,
    db: Session = Depends(get_db)
):
    query = db.query(ProjectExperience)

    if search:
        query = query.filter(ProjectExperience.project_name.ilike(f"%{search}%"))

    total = query.count()
    data = query.offset(skip).limit(limit).all()

    return PaginatedResponseSchemas.PaginatedResponse(data=data, total=total)

@router.get("/export/xlsx")
def export_to_xlsx(
    search: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """Export all project experiences to XLSX file"""
    
    # Query all data
    query = (
        db.query(ProjectExperience)
        .options(joinedload(ProjectExperience.consulting_manager))
    )
    
    if search:
        query = query.filter(ProjectExperience.project_name.ilike(f"%{search}%"))
    
    results = query.all()
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Project Experiences"
    
    # Define headers
    headers = [
        "ID",
        "No Sales Order",
        "Customer Name",
        "Project Name",
        "Project Year",
        "Category",
        "Consulting Manager"
    ]
    
    # Style for headers
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Write headers
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Write data
    for row_idx, record in enumerate(results, start=2):
        ws.cell(row=row_idx, column=1, value=record.id)
        ws.cell(row=row_idx, column=2, value=record.no_sales_order)
        ws.cell(row=row_idx, column=3, value=record.customer_name)
        ws.cell(row=row_idx, column=4, value=record.project_name)
        ws.cell(row=row_idx, column=5, value=record.project_year)
        ws.cell(row=row_idx, column=6, value=record.category)
        ws.cell(row=row_idx, column=7, value=record.consulting_manager.name if record.consulting_manager else "")
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Generate filename with timestamp
    filename = f"project_experiences_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    # Return as streaming response
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@router.put("/{project_id}", response_model=ProjectExperienceResponse)
def update_project_experience(
    project_id: int,
    data: ProjectExperienceUpdate,
    db: Session = Depends(get_db)
):
    project = db.query(ProjectExperience).filter_by(id=project_id).first()
    if not project:
        raise HTTPException(404, "Project experience not found")

    if data.consulting_manager_id:
        manager = db.query(ConsultingManager).filter_by(id=data.consulting_manager_id).first()
        if not manager:
            raise HTTPException(404, "Consulting manager does not exist")

        project.consulting_manager_id = data.consulting_manager_id

    db.commit()
    db.refresh(project)
    return project


# Delete project
@router.delete("/{project_id}")
def delete_experience(project_id: int, db: Session = Depends(get_db)):
    proj = db.query(ProjectExperience).filter_by(id=project_id).first()
    if not proj:
        raise HTTPException(404, "Project experience not found")

    db.delete(proj)
    db.commit()
    return {"message": "Deleted successfully"}
